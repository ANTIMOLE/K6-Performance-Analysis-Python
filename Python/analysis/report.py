"""
report.py — ZENIT Performance Testing
Generate laporan: Excel (.xlsx) + chart PNG.

FIXES (cumulative):
  BUG-02:     _get_chart_vals() pakai actual rest_vals/trpc_vals dari analysis dict
              (disimpan oleh analyze_load/analyze_exploratory di stats.py).
              Sebelumnya: _reconstruct_vals approximation menyebabkan spread 2x
              terlalu sempit di boxplot dan paired scatter.
  BUG-03-FMT: _fmt_float() handle inf/nan → "N/A", bukan crash atau #NUM! di Excel.
  BUG-04-SOAK: Soak timeseries x-axis: setiap side pakai local var, tambah note
              bahwa REST dan tRPC berjalan independen (tidak tumpuk di waktu sama).
  ISS-04:     Forest plot hapus CI fallback ±0.3 arbitrary. Jika CI tidak tersedia,
              tampilkan titik saja tanpa error bar.
  ISS-05:     Excel coloring direction-aware: metrik "lower is better" (latency,
              cpu, ram, error) → warna biru jika REST lebih rendah (REST menang).
              Metrik "higher is better" (throughput) → warna biru jika REST lebih tinggi.
  ISS-06:     Raw JSON: strip difference_scores dari endpoint_metrics juga.
  n_warning:  Propagate n_warning ke Excel Summary sheet.

REMOVED:
  bab5_draft.md export dan sheet Excel Narasi/Research_Questions dihapus.
  Auto-generated narrative tidak dipakai sebagai sumber teks Bab 5 dan
  interpretasi RQ1/RQ2/RQ3 tetap tampil di terminal (main.py), bukan di
  file laporan. Excel sekarang berisi data sheets saja: Summary, detail
  per group, Soak, Decomposition, CrossScenario, EffectSizeMatrix,
  SoakComparison, OrderEffects, ChartDesc.
"""

import json
import math
import os

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# KONSTANTA
# ---------------------------------------------------------------------------

COLORS = {
    "rest":      "#2196F3",
    "trpc":      "#FF5722",
    "header_bg": "FF1565C0",
    "header_fg": "FFFFFFFF",
    "sig_rest":  "FFE3F2FD",   # Biru muda — REST lebih baik (signifikan)
    "sig_trpc":  "FFFFF3E0",   # Oranye muda — tRPC lebih baik (signifikan)
    "sig_pos":   "FFE8F5E9",   # Hijau muda — legacy positif
    "ma_rest":   "#0D47A1",
    "ma_trpc":   "#BF360C",
    "threshold": "#F44336",
}

METRIC_LABELS = {
    "avg_rt":             "Avg RT (ms)",
    "med_rt":             "Median RT (ms)",
    "p90":                "P90 (ms)",
    "p95":                "P95 (ms)",
    "p99":                "P99 (ms)",
    "min_rt":             "Min RT (ms)",
    "max_rt":             "Max RT (ms)",
    "throughput":         "Throughput (req/s)",
    "functional_error":   "Func. Error Rate",
    "http_req_failed":    "HTTP Req Failed Rate",
    "sla_breach":         "SLA Breach Rate",
    "checks_pass_rate":   "Checks Pass Rate",
    "payload_bytes":      "Payload Avg (bytes)",
    "http_count":         "HTTP Req Count",
    "cpu_pct":            "CPU % (backend)",
    "cpu_total_pct":      "System CPU (%)",
    "mem_mb":             "RAM MB (backend)",
    "pg_active":          "DB Connections",
    "pg_cache_hit_ratio": "DB Cache Hit %",
    "pg_tps_delta":       "DB TPS Delta",
    "db_query_avg_ms":    "DB Query Avg (ms)",
    "network_total_kb_s": "Network I/O (KB/s)",
}

# ISS-05: Metrik yang "lower is better" — REST menang jika diff < 0 (REST lebih rendah)
# BUG-C FIX: payload_bytes ditambahkan — tRPC cenderung punya payload lebih besar
# (JSON envelope overhead), sehingga REST dengan payload lebih kecil = menang.
LOWER_IS_BETTER = {
    "avg_rt", "med_rt", "p90", "p95", "p99", "min_rt", "max_rt",
    "cpu_pct", "cpu_total_pct", "mem_mb", "functional_error", "http_req_failed", "sla_breach", "db_query_avg_ms",
    "payload_bytes",
}

INFERENTIAL_METRICS = ["avg_rt", "p95", "p99", "throughput", "cpu_pct", "mem_mb"]

# HYPOTHESIS_MAPPING dihapus — thesis menggunakan RQ-based framing (RQ1/RQ2/RQ3).
# Sheet Hypotheses dan Research_Questions tidak diekspor — RQ tetap ditampilkan
# di terminal (main.py) tapi tidak masuk Excel maupun draft markdown.


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def _fmt_float(v, digits=3):
    """
    BUG-03-FMT FIX: Handle None, inf, nan → "N/A" (bukan crash atau #NUM! Excel).
    """
    if v is None:
        return "N/A"
    try:
        f = float(v)
        if math.isnan(f) or math.isinf(f):
            return "N/A"
        return round(f, digits)
    except (TypeError, ValueError):
        return "N/A"


def _header_style(ws, row, col, value, bold=True, bg=None, fg="000000"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=fg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


def _excel_bg(metric_name: str, mean_diff, sig, cd_val) -> str | None:
    """
    ISS-05 FIX: Direction-aware Excel coloring.
    - Lower-is-better metrics: REST menang jika mean_diff < 0 (REST lebih rendah)
    - Higher-is-better metrics: REST menang jika mean_diff > 0 (REST lebih tinggi)
    Warna:
    - BIRU MUDA (sig_rest): REST menang, signifikan
    - ORANYE MUDA (sig_trpc): tRPC menang, signifikan
    - Tidak sig atau efek trivial: tidak ada warna
    """
    if not sig or cd_val is None or abs(cd_val) < 0.2:
        return None
    if mean_diff is None:
        return None

    lower_is_better = metric_name in LOWER_IS_BETTER
    if lower_is_better:
        rest_wins = mean_diff < 0  # REST lebih rendah → REST menang
    else:
        rest_wins = mean_diff > 0  # REST lebih tinggi → REST menang

    return COLORS["sig_rest"] if rest_wins else COLORS["sig_trpc"]


def _get_mean(group, metric_name, side):
    m = group["metrics"].get(metric_name, {})
    return (m.get("descriptive") or {}).get(side, {}).get("mean")


def _save(fig, output_dir, fname):
    fpath = os.path.join(output_dir, fname)
    plt.savefig(fpath, dpi=150, bbox_inches="tight")
    plt.close()
    return fpath


def _get_chart_vals(analysis: dict):
    """
    BUG-02 FIX: Ambil actual per-run REST/tRPC values dari analysis dict.
    stats.py sekarang menyimpan rest_vals dan trpc_vals di setiap analysis dict.
    Jika tidak ada (fallback untuk soak/observational), return None, None.
    """
    rest_vals = analysis.get("rest_vals")
    trpc_vals = analysis.get("trpc_vals")
    if rest_vals and trpc_vals and len(rest_vals) == len(trpc_vals):
        return list(rest_vals), list(trpc_vals)
    return None, None


# ---------------------------------------------------------------------------
# EXCEL SHEETS
# ---------------------------------------------------------------------------

def write_summary_sheet(wb, all_results: dict):
    ws = wb.create_sheet("Summary")

    headers = [
        "Scenario", "Test Type", "Condition", "N", "N Warning", "Metrik",
        "REST Mean", "REST SD", "tRPC Mean", "tRPC SD",
        "Mean Diff", "Diff%", "Cohen's d", "Magnitude",
        "SW p", "Normal?", "Test Used", "p-value", "Sig?",
        "CI Lower", "CI Upper", "CI Covers 0?", "Kesimpulan",
    ]
    for ci, h in enumerate(headers, 1):
        _header_style(ws, 1, ci, h, bg=COLORS["header_bg"], fg=COLORS["header_fg"])

    row = 2
    for group_key, group in sorted(all_results["groups"].items()):
        scenario  = group["scenario"]
        test_type = group["test_type"]
        condition = group.get("condition", "C2")
        n         = group["n"]

        for metric_name, analysis in group["metrics"].items():
            if analysis.get("status") in ("no_data", "insufficient_data"):
                continue

            desc    = analysis.get("descriptive", {})
            dr      = desc.get("rest", {})
            dt      = desc.get("trpc", {})
            cd      = analysis.get("cohens_d", {})
            bci     = analysis.get("bootstrap_ci", {})
            inf     = analysis.get("inferential", analysis.get("ttest_reference", {}))
            sw      = analysis.get("shapiro_wilk", {})
            test_u  = analysis.get("test_used", "ref")
            n_warn  = analysis.get("n_warning", "")

            rm        = dr.get("mean")
            tm        = dt.get("mean")
            mean_diff = (rm - tm) if rm is not None and tm is not None else None
            diff_pct  = (mean_diff / tm * 100) if mean_diff is not None and tm and tm != 0 else None
            sig       = inf.get("significant")
            cd_val    = cd.get("d")
            cd_mag    = cd.get("magnitude", "")

            # ISS-05 FIX: direction-aware coloring
            bg = _excel_bg(metric_name, mean_diff, sig, cd_val)

            # Issue #3 Fix: ⚠ jika d tidak_interpretatif
            if cd_mag == "tidak_interpretatif":
                sig_display = "⚠ d≠"
            else:
                sig_display = "Ya" if sig else ("Tidak" if sig is False else "N/A")

            row_vals = [
                scenario, test_type, condition, n,
                n_warn[:60] if n_warn else "",          # n_warning propagated
                METRIC_LABELS.get(metric_name, metric_name),
                _fmt_float(rm), _fmt_float(dr.get("std")),
                _fmt_float(tm), _fmt_float(dt.get("std")),
                _fmt_float(mean_diff),
                f"{diff_pct:+.1f}%" if diff_pct is not None else "N/A",
                _fmt_float(cd_val), cd.get("magnitude", "N/A"),
                _fmt_float(sw.get("p"), 4) if sw.get("p") is not None else "N/A",
                "Ya" if sw.get("normal") else ("Tidak" if sw.get("normal") is False else "N/A"),
                test_u,
                _fmt_float(inf.get("p"), 4) if inf.get("p") is not None else "N/A",
                sig_display,
                _fmt_float(bci.get("ci_lower")),
                _fmt_float(bci.get("ci_upper")),
                "Ya" if bci.get("covers_zero") else ("Tidak" if bci.get("covers_zero") is False else "N/A"),
                analysis.get("conclusion", analysis.get("framing", ""))[:60],
            ]

            for ci, val in enumerate(row_vals, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if bg:
                    cell.fill = PatternFill("solid", fgColor=bg)
            row += 1

    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["W"].width = 40
    ws.freeze_panes = "A2"


def write_detail_sheet(wb, group_key: str, group: dict):
    condition = group.get("condition", "C2")
    title = f"{group['scenario']}_{group['test_type']}_{condition}"
    ws = wb.create_sheet(title[:30])

    ws.merge_cells("A1:P1")
    ws["A1"] = (f"{group['scenario'].upper()} — {group['test_type'].upper()} "
                f"/ {condition} (N={group['n']})")
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=COLORS["header_bg"])
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    headers = [
        "Metrik", "REST Mean", "REST SD", "tRPC Mean", "tRPC SD",
        "Mean Diff", "Diff%", "Cohen's d", "Magnitude",
        "SW p", "Normal?", "Test", "p-value", "Sig?",
        "Kesimpulan", "N Warning",
    ]
    for ci, h in enumerate(headers, 1):
        _header_style(ws, row, ci, h, bg="FF1565C0", fg="FFFFFFFF")
    row += 1

    for metric_name, analysis in group["metrics"].items():
        if analysis.get("status") in ("no_data", "insufficient_data"):
            continue

        desc   = analysis.get("descriptive", {})
        dr     = desc.get("rest", {})
        dt     = desc.get("trpc", {})
        cd     = analysis.get("cohens_d", {})
        inf    = analysis.get("inferential", analysis.get("ttest_reference", {}))
        sw     = analysis.get("shapiro_wilk", {})
        test_u = analysis.get("test_used", "ref")
        n_warn = analysis.get("n_warning", "")

        rm     = dr.get("mean")
        tm     = dt.get("mean")
        diff   = (rm - tm) if rm is not None and tm is not None else None
        diffp  = (diff / tm * 100) if diff is not None and tm and tm != 0 else None
        sig    = inf.get("significant")
        cd_val = cd.get("d")
        cd_mag = cd.get("magnitude", "")

        # ISS-05 FIX: direction-aware coloring
        bg = _excel_bg(metric_name, diff, sig, cd_val)

        # Issue #3 Fix: ⚠ jika d tidak_interpretatif meski sig=True
        if cd_mag == "tidak_interpretatif":
            sig_display = "⚠ d≠"
        else:
            sig_display = "✓" if sig else ("✗" if sig is False else "—")

        row_data = [
            METRIC_LABELS.get(metric_name, metric_name),
            _fmt_float(rm), _fmt_float(dr.get("std")),
            _fmt_float(tm), _fmt_float(dt.get("std")),
            _fmt_float(diff),
            f"{diffp:+.1f}%" if diffp is not None else "N/A",
            _fmt_float(cd_val), cd.get("magnitude", "N/A"),
            _fmt_float(sw.get("p"), 4) if sw.get("p") is not None else "N/A",
            "Ya" if sw.get("normal") else ("Tidak" if sw.get("normal") is False else "N/A"),
            test_u,
            _fmt_float(inf.get("p"), 4) if inf.get("p") is not None else "N/A",
            sig_display,
            analysis.get("conclusion", "—")[:50],
            n_warn[:50] if n_warn else "",
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
        row += 1

    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["O"].width = 40
    ws.column_dimensions["P"].width = 40
    ws.freeze_panes = "A4"



def write_soak_sheet(wb, group_key: str, group: dict):
    title = f"Soak_{group['scenario']}"
    ws = wb.create_sheet(title[:30])

    ws.merge_cells("A1:H1")
    ws["A1"] = f"SOAK — {group['scenario'].upper()} (Observational, N=1)"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=COLORS["header_bg"])
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    ws.cell(row=row, column=1, value="RINGKASAN DESKRIPTIF").font = Font(bold=True)
    row += 1
    for lbl, col in [("Metrik",1),("REST",2),("tRPC",3),("Δ",4),("Δ %",5)]:
        _header_style(ws, row, col, lbl, bg="FF1565C0", fg="FFFFFFFF")
    row += 1

    for metric_name, analysis in group["metrics"].items():
        if analysis.get("status") in ("no_data", "insufficient_data"):
            continue
        desc = analysis.get("descriptive", {})
        rm   = desc.get("rest", {}).get("mean")
        tm   = desc.get("trpc", {}).get("mean")
        diff = (rm - tm) if rm is not None and tm is not None else None
        diffp = (diff / tm * 100) if diff is not None and tm and tm != 0 else None

        ws.cell(row=row, column=1, value=METRIC_LABELS.get(metric_name, metric_name))
        ws.cell(row=row, column=2, value=_fmt_float(rm))
        ws.cell(row=row, column=3, value=_fmt_float(tm))
        ws.cell(row=row, column=4, value=_fmt_float(diff))
        ws.cell(row=row, column=5, value=f"{diffp:+.1f}%" if diffp is not None else "N/A")
        row += 1

    ts      = group.get("timeseries", {}) or {}
    rest_ts = ts.get("rest") or {}
    trpc_ts = ts.get("trpc") or {}

    row += 1
    ws.cell(row=row, column=1, value="MEMORY LEAK ANALYSIS (Linear Regression)").font = Font(bold=True)
    row += 1
    for side, ts_data in [("REST", rest_ts), ("tRPC", trpc_ts)]:
        note     = ts_data.get("mem_slope_note", "Data tidak tersedia")
        slope_mh = ts_data.get("mem_slope_mb_per_hour")
        interval = ts_data.get("sampling_interval_sec", 5)
        ws.cell(row=row, column=1, value=side)
        ws.cell(row=row, column=2, value=note)
        if slope_mh is not None:
            r2       = ts_data.get("mem_slope_r2")
            ws.cell(row=row, column=3, value=f"{slope_mh:.2f} MB/hour (@ {interval:.0f}s)")
            # BUG-03 FIX: >= bukan > (slope==10 dulu jatuh ke "TURUN")
            # BUG-04 FIX: seragamkan logic dengan write_soak_comparison_sheet
            if r2 is not None and r2 >= 0.5 and slope_mh > 5:
                interp = "⚠ POTENSI LEAK (R²≥0.5, slope>5)"
            elif slope_mh > 10:
                interp = "NAIK SIGNIFIKAN"
            elif slope_mh > 3:
                interp = "NAIK MODERAT"
            elif slope_mh < -3:
                interp = "TURUN"
            else:
                interp = "STABIL"
            ws.cell(row=row, column=4, value=interp)
        row += 1

    for ci in range(1, 9):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.column_dimensions["B"].width = 65


def write_decomposition_sheet(wb, decomp_results: dict):
    ws = wb.create_sheet("Decomposition")
    ws.merge_cells("A1:K1")
    ws["A1"] = "DECOMPOSITION ANALYSIS — C3 (Auth Overhead) & C4 (Batching Benefit)"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=COLORS["header_bg"])
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    ws.cell(row=row, column=1, value="C3 — Auth Overhead Contribution").font = Font(bold=True, size=11)
    row += 1
    headers_c3 = ["Scenario","Metrik","C1 (REST)","C2 (tRPC base)","C3 (tRPC+auth)",
                   "Gap C1-C2","Gap C1-C3","Auth Contrib %","Remaining %","Catatan"]
    for ci, h in enumerate(headers_c3, 1):
        _header_style(ws, row, ci, h, bg="FF1565C0", fg="FFFFFFFF")
    row += 1

    for gk, dc in decomp_results.items():
        if dc.get("condition") != "C3":
            continue
        for metric_name, mdata in dc.get("metrics", {}).items():
            if mdata.get("status") == "missing_data":
                ws.cell(row=row, column=1, value=dc.get("scenario"))
                ws.cell(row=row, column=2, value=METRIC_LABELS.get(metric_name, metric_name))
                ws.cell(row=row, column=10, value=mdata.get("note", ""))
                row += 1
                continue

            # Issue #2 Fix: anomaly flag coloring
            anomaly = mdata.get("anomaly_flag", False)
            anomaly_bg = "FFFFF3E0" if anomaly else None  # oranye muda jika anomaly

            row_vals = [
                dc.get("scenario"),
                METRIC_LABELS.get(metric_name, metric_name),
                _fmt_float(mdata.get("c1_mean")),
                _fmt_float(mdata.get("c2_mean")),
                _fmt_float(mdata.get("c3_mean")),
                _fmt_float(mdata.get("gap_c1c2")),
                _fmt_float(mdata.get("gap_c1c3")),
                _fmt_float(mdata.get("auth_contribution_pct"), 1),
                _fmt_float(mdata.get("remaining_gap_pct"), 1),
                # Issue #2 Fix: tampilkan anomaly_note jika ada, fallback ke note biasa
                (mdata.get("anomaly_note") or mdata.get("note") or "")[:120],
            ]
            for ci, val in enumerate(row_vals, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.alignment = Alignment(horizontal="center")
                if anomaly_bg:
                    cell.fill = PatternFill("solid", fgColor=anomaly_bg)
            row += 1

    row += 1
    ws.cell(row=row, column=1, value="C4 — Batching Benefit").font = Font(bold=True, size=11)
    row += 1
    headers_c4 = ["Scenario","C2 HTTP","C4 HTTP","HTTP Red%",
                   "C2 P95","C4 P95","Lat Red%","C2 Tput","C4 Tput","Tput Gain%","Catatan"]
    for ci, h in enumerate(headers_c4, 1):
        _header_style(ws, row, ci, h, bg="FF1565C0", fg="FFFFFFFF")
    row += 1

    for gk, dc in decomp_results.items():
        if dc.get("condition") != "C4":
            continue
        cd4 = dc.get("decomposition", {})
        if cd4.get("status") == "missing_data":
            ws.cell(row=row, column=1, value=dc.get("scenario"))
            ws.cell(row=row, column=11, value=cd4.get("note", ""))
            row += 1
            continue
        row_vals = [
            dc.get("scenario"),
            _fmt_float(cd4.get("c2_http_count"), 0),
            _fmt_float(cd4.get("c4_http_count"), 0),
            _fmt_float(cd4.get("http_reduction_pct"), 1),
            _fmt_float(cd4.get("c2_p95")),
            _fmt_float(cd4.get("c4_p95")),
            _fmt_float(cd4.get("latency_reduction_pct"), 1),
            _fmt_float(cd4.get("c2_throughput")),
            _fmt_float(cd4.get("c4_throughput")),
            _fmt_float(cd4.get("throughput_gain_pct"), 1),
            (cd4.get("note") or "")[:80],
        ]
        for ci, val in enumerate(row_vals, 1):
            ws.cell(row=row, column=ci, value=val).alignment = Alignment(horizontal="center")
        row += 1

    for ci in range(1, 12):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["K"].width = 50


# ---------------------------------------------------------------------------
# NEW SHEETS (Fix: CrossScenario, EffectSizeMatrix, SoakComparison, OrderEffects)
# ---------------------------------------------------------------------------

# Urutan skenario untuk konsistensi tampilan
SCENARIO_ORDER  = ["s01_browse", "s02_shopping", "s03_checkout", "s04_auth", "s05_admin"]
SCENARIO_LABELS = {
    "s01_browse":   "S01 Browse",
    "s02_shopping": "S02 Shopping",
    "s03_checkout": "S03 Checkout",
    "s04_auth":     "S04 Auth",
    "s05_admin":    "S05 Admin",
}

# Metrik yang ditampilkan di CrossScenario & EffectSizeMatrix
CROSS_METRICS = ["p95", "throughput", "avg_rt", "cpu_pct", "mem_mb"]

# Magnitude → warna background untuk EffectSizeMatrix
_MAG_COLOR = {
    "besar":             "FFB2EBF2",   # cyan muda
    "sedang":            "FFFFF9C4",   # kuning muda
    "kecil":             "FFF1F8E9",   # hijau sangat muda
    "trivial":           "FFF5F5F5",   # abu sangat muda
    "tidak_interpretatif": "FFFFF3E0", # oranye muda
}


def write_cross_scenario_sheet(wb, all_results: dict):
    """
    CrossScenario sheet: perbandingan semua skenario Load C2 dalam satu view.
    Sangat berguna untuk nulis bab 4 — tidak perlu lompat-lompat antar sheet.

    Layout:
      Baris  : setiap skenario (S01–S05)
      Kolom  : per metrik → REST Mean | tRPC Mean | Δ | d | Magnitude | Sig?
    """
    ws = wb.create_sheet("CrossScenario")

    ws.merge_cells("A1:Z1")
    ws["A1"] = "CROSS-SCENARIO COMPARISON — Load Test C2 (N=10)"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=COLORS["header_bg"])
    ws["A1"].alignment = Alignment(horizontal="center")

    groups = all_results.get("groups", {})

    # --- Header baris 2 & 3 (merged per metrik) ---
    row_hdr1 = 3
    row_hdr2 = 4
    col = 2  # kolom A untuk label skenario

    ws.cell(row=row_hdr1, column=1, value="Scenario").font = Font(bold=True)
    ws.cell(row=row_hdr2, column=1, value="").font         = Font(bold=True)

    metric_col_start = {}
    for mn in CROSS_METRICS:
        label    = METRIC_LABELS.get(mn, mn)
        # Merge 5 kolom per metrik di row_hdr1
        ws.merge_cells(
            start_row=row_hdr1, start_column=col,
            end_row=row_hdr1,   end_column=col + 4
        )
        c = ws.cell(row=row_hdr1, column=col, value=label)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=COLORS["header_bg"])
        c.alignment = Alignment(horizontal="center")

        metric_col_start[mn] = col
        sub_hdrs = ["REST", "tRPC", "Δ", "d", "Sig?"]
        for j, sh in enumerate(sub_hdrs):
            sc = ws.cell(row=row_hdr2, column=col + j, value=sh)
            sc.font      = Font(bold=True)
            sc.fill      = PatternFill("solid", fgColor="FFE3F2FD")
            sc.alignment = Alignment(horizontal="center")
        col += 5

    # --- Data rows ---
    data_row = row_hdr2 + 1   # FIX: satu counter di luar loop — tidak reset per test_type

    for test_type in ("load", "stress", "spike"):
        wrote_section_header = False

        for sc_name in SCENARIO_ORDER:
            gk    = f"{sc_name}__{test_type}__C2"
            group = groups.get(gk)
            if group is None:
                continue

            if not wrote_section_header:
                # Section divider
                ws.merge_cells(f"A{data_row}:Z{data_row}")
                sec_cell = ws.cell(row=data_row, column=1,
                                   value=f"── {test_type.upper()} (N={'10' if test_type=='load' else '3'}) ──")
                sec_cell.font = Font(bold=True, italic=True, color="FFFFFF")
                sec_cell.fill = PatternFill("solid", fgColor="FF455A64")
                sec_cell.alignment = Alignment(horizontal="left")
                data_row += 1
                wrote_section_header = True

            # Skenario label
            lbl_cell = ws.cell(row=data_row, column=1,
                               value=SCENARIO_LABELS.get(sc_name, sc_name))
            lbl_cell.font      = Font(bold=True)
            lbl_cell.alignment = Alignment(horizontal="left")

            is_expl = test_type in ("stress", "spike")

            for mn in CROSS_METRICS:
                c0 = metric_col_start[mn]
                a  = group["metrics"].get(mn, {})
                if a.get("status") in ("no_data", "insufficient_data"):
                    for j in range(5):
                        ws.cell(row=data_row, column=c0 + j, value="N/A").alignment = Alignment(horizontal="center")
                    continue

                desc    = a.get("descriptive", {})
                rm      = desc.get("rest",  {}).get("mean")
                tm      = desc.get("trpc",  {}).get("mean")
                diff    = (rm - tm) if rm is not None and tm is not None else None
                cd      = a.get("cohens_d", {})
                d_val   = cd.get("d")
                cd_mag  = cd.get("magnitude", "")
                inf     = a.get("inferential", a.get("ttest_reference", {}))
                sig     = inf.get("significant")

                # Sig display — Issue #3 aware
                if cd_mag == "tidak_interpretatif":
                    sig_str = "⚠"
                elif sig is True:
                    sig_str = "✓*" if is_expl else "✓"
                elif sig is False:
                    sig_str = "✗"
                else:
                    sig_str = "—"

                vals = [
                    _fmt_float(rm, 2),
                    _fmt_float(tm, 2),
                    _fmt_float(diff, 2) if diff is not None else "N/A",
                    _fmt_float(d_val, 2) if d_val is not None else "N/A",
                    sig_str,
                ]

                # Warna background berdasarkan magnitude d
                bg_hex = _MAG_COLOR.get(cd_mag)

                for j, val in enumerate(vals):
                    cell = ws.cell(row=data_row, column=c0 + j, value=val)
                    cell.alignment = Alignment(horizontal="center")
                    if bg_hex and j < 4:  # jangan warnai kolom Sig?
                        cell.fill = PatternFill("solid", fgColor=bg_hex)

                # Sig? kolom: warna khusus jika signifikan
                sig_cell = ws.cell(row=data_row, column=c0 + 4)
                if sig is True and cd_mag != "tidak_interpretatif":
                    rest_wins = (diff or 0) < 0 if mn in LOWER_IS_BETTER else (diff or 0) > 0
                    sig_cell.fill = PatternFill("solid", fgColor=COLORS["sig_rest"] if rest_wins else COLORS["sig_trpc"])

            data_row += 1

    # Kolom widths
    ws.column_dimensions["A"].width = 16
    for mn in CROSS_METRICS:
        c0 = metric_col_start[mn]
        for j in range(5):
            ws.column_dimensions[get_column_letter(c0 + j)].width = 9
    ws.freeze_panes = "B5"

    # Legend
    legend_row = data_row + 2
    ws.cell(row=legend_row, column=1, value="Legenda warna d:").font = Font(bold=True)
    legend_items = [("Besar |d|≥0.8", "FFB2EBF2"), ("Sedang 0.5≤|d|<0.8", "FFFFF9C4"),
                    ("Kecil 0.2≤|d|<0.5", "FFF1F8E9"), ("Trivial |d|<0.2", "FFF5F5F5"),
                    ("d tidak interpretatif", "FFFFF3E0")]
    for j, (lbl, col_hex) in enumerate(legend_items):
        c = ws.cell(row=legend_row, column=j + 2, value=lbl)
        c.fill = PatternFill("solid", fgColor=col_hex)
        c.alignment = Alignment(horizontal="center")


def write_effect_size_matrix_sheet(wb, all_results: dict):
    """
    EffectSizeMatrix sheet: tabel d-value per (metrik × skenario) untuk Load C2.
    Langsung bisa di-copy ke tabel skripsi bab 4 sebagai "Effect Size Consistency Table".

    Baris  : metrik
    Kolom  : skenario S01–S05
    Nilai  : Cohen's d (negatif = tRPC lebih tinggi, positif = REST lebih tinggi)
    """
    ws = wb.create_sheet("EffectSizeMatrix")

    ws.merge_cells("A1:H1")
    ws["A1"] = "EFFECT SIZE MATRIX — Cohen's d per (Metrik × Skenario) — Load C2 (N=10)"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=COLORS["header_bg"])
    ws["A1"].alignment = Alignment(horizontal="center")

    groups = all_results.get("groups", {})

    # Header: Metrik | S01 | S02 | S03 | S04 | S05 | Pattern
    headers = ["Metrik"] + [SCENARIO_LABELS[sc] for sc in SCENARIO_ORDER] + ["Pattern"]
    row = 3
    for ci, h in enumerate(headers, 1):
        _header_style(ws, row, ci, h, bg=COLORS["header_bg"], fg=COLORS["header_fg"])

    # Sub-header untuk d dan Sig
    row_sub = 4
    ws.cell(row=row_sub, column=1, value="(d negatif = tRPC lebih tinggi)").font = Font(italic=True, size=9)

    row = 5
    all_metrics_for_matrix = [
        "p95", "p99", "avg_rt", "throughput",
        "cpu_pct", "mem_mb", "db_query_avg_ms", "network_total_kb_s",
        "payload_bytes", "sla_breach", "functional_error",
    ]

    for mn in all_metrics_for_matrix:
        ws.cell(row=row, column=1, value=METRIC_LABELS.get(mn, mn)).font = Font(bold=True)

        d_vals = []
        for ci, sc_name in enumerate(SCENARIO_ORDER, 2):
            gk    = f"{sc_name}__load__C2"
            group = groups.get(gk)
            if group is None:
                ws.cell(row=row, column=ci, value="—").alignment = Alignment(horizontal="center")
                d_vals.append(None)
                continue

            a      = group["metrics"].get(mn, {})
            cd     = a.get("cohens_d", {})
            d_val  = cd.get("d")
            cd_mag = cd.get("magnitude", "")
            inf    = a.get("inferential", {})
            sig    = inf.get("significant")

            if d_val is None or a.get("status") in ("no_data", "insufficient_data"):
                ws.cell(row=row, column=ci, value="N/A").alignment = Alignment(horizontal="center")
                d_vals.append(None)
                continue

            cell = ws.cell(row=row, column=ci, value=round(float(d_val), 2))
            cell.alignment = Alignment(horizontal="center")

            # Warna background berdasarkan magnitude
            bg_hex = _MAG_COLOR.get(cd_mag)
            if bg_hex:
                cell.fill = PatternFill("solid", fgColor=bg_hex)

            # Bold jika signifikan
            if sig:
                cell.font = Font(bold=True)

            d_vals.append(d_val)

        # Pattern kolom: deteksi arah konsisten
        # Gunakan arah d tanpa threshold magnitude — cukup hitung wins per direction
        # Untuk semua metrik: d < 0 artinya REST < tRPC
        #   lower_is_better (p95, rt, sla): d<0 = REST menang
        #   higher_is_better (throughput, http): d<0 = tRPC menang
        #   resource (cpu, mem): d<0 = REST lebih efisien = REST menang
        LOWER_IB = {"p95","p99","avg_rt","sla_breach","functional_error","db_query_avg_ms"}
        HIGHER_IB = {"throughput","http_count"}

        valid_d = [d for d in d_vals if d is not None]
        if len(valid_d) >= 3:
            if mn in LOWER_IB or mn not in HIGHER_IB:
                # d<0 = REST wins (lower value = better, or resource)
                rest_w = sum(1 for d in valid_d if d < 0)
                trpc_w = sum(1 for d in valid_d if d > 0)
            else:
                # d>0 = REST wins (higher value = better)
                rest_w = sum(1 for d in valid_d if d > 0)
                trpc_w = sum(1 for d in valid_d if d < 0)

            if rest_w >= 4:
                pattern = f"REST lebih baik ({rest_w}/5 skenario)"
            elif trpc_w >= 4:
                pattern = f"tRPC lebih baik ({trpc_w}/5 skenario)"
            elif rest_w >= 3:
                pattern = f"REST cenderung lebih baik ({rest_w}/5)"
            elif trpc_w >= 3:
                pattern = f"tRPC cenderung lebih baik ({trpc_w}/5)"
            else:
                pattern = "Mixed/tidak konsisten"
        else:
            pattern = "—"

        ws.cell(row=row, column=len(SCENARIO_ORDER) + 2, value=pattern).alignment = Alignment(horizontal="left")
        row += 1

    # Widths
    ws.column_dimensions["A"].width = 22
    for i in range(2, len(SCENARIO_ORDER) + 3):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.freeze_panes = "B5"

    # Legend
    row += 1
    ws.cell(row=row, column=1, value="Legenda: Bold = signifikan (p<0.05). Warna = magnitude Cohen's d.").font = Font(italic=True, size=9)
    row += 1
    legend_items = [("Besar |d|≥0.8", "FFB2EBF2"), ("Sedang", "FFFFF9C4"),
                    ("Kecil", "FFF1F8E9"), ("Trivial", "FFF5F5F5"), ("d tidak interpretatif", "FFFFF3E0")]
    for j, (lbl, col_hex) in enumerate(legend_items):
        c = ws.cell(row=row, column=j + 1, value=lbl)
        c.fill = PatternFill("solid", fgColor=col_hex)
        c.alignment = Alignment(horizontal="center")
        c.font = Font(size=9)


def write_soak_comparison_sheet(wb, all_results: dict):
    """
    SoakComparison sheet: ringkasan memory slope semua skenario soak dalam satu tabel.
    Berguna untuk pembahasan S04 memory leak (R²=0.606) vs skenario lain.
    """
    ws = wb.create_sheet("SoakComparison")

    ws.merge_cells("A1:J1")
    ws["A1"] = "SOAK MEMORY SLOPE COMPARISON — Semua Skenario"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=COLORS["header_bg"])
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "Scenario",
        "REST Slope (MB/hr)", "REST R²", "REST p", "REST Interpretasi",
        "tRPC Slope (MB/hr)", "tRPC R²", "tRPC p", "tRPC Interpretasi",
        "Catatan",
    ]
    row = 3
    for ci, h in enumerate(headers, 1):
        _header_style(ws, row, ci, h, bg=COLORS["header_bg"], fg=COLORS["header_fg"])
    row += 1

    groups = all_results.get("groups", {})

    for sc_name in SCENARIO_ORDER:
        gk    = f"{sc_name}__soak__C2"
        group = groups.get(gk)
        if group is None:
            continue

        ts      = group.get("timeseries") or {}
        rest_ts = ts.get("rest") or {}
        trpc_ts = ts.get("trpc") or {}

        def _interp(slope_mh, r2):
            if slope_mh is None:
                return "N/A"
            if r2 is not None and r2 >= 0.5 and slope_mh > 5:
                return "⚠ POTENSI LEAK (R²≥0.5, slope>5)"
            elif slope_mh > 10:
                return "NAIK SIGNIFIKAN"
            elif slope_mh > 3:
                return "NAIK MODERAT"
            elif slope_mh < -3:
                return "TURUN"
            else:
                return "STABIL"

        def _r2_note(r2_rest, r2_trpc, slope_rest, slope_trpc):
            notes = []
            if r2_rest is not None and r2_rest >= 0.4:
                notes.append(f"REST R²={r2_rest:.3f} — trend linear kuat")
            if r2_trpc is not None and r2_trpc >= 0.4:
                notes.append(f"tRPC R²={r2_trpc:.3f} — trend linear kuat")
            return " | ".join(notes) if notes else "Trend lemah/tidak linear"

        r_slope = rest_ts.get("mem_slope_mb_per_hour")
        t_slope = trpc_ts.get("mem_slope_mb_per_hour")
        r_r2    = rest_ts.get("mem_slope_r2")
        t_r2    = trpc_ts.get("mem_slope_r2")
        r_p     = rest_ts.get("mem_slope_p")
        t_p     = trpc_ts.get("mem_slope_p")

        row_vals = [
            SCENARIO_LABELS.get(sc_name, sc_name),
            _fmt_float(r_slope, 2) if r_slope is not None else "N/A",
            _fmt_float(r_r2, 3) if r_r2 is not None else "N/A",
            _fmt_float(r_p, 4) if r_p is not None else "N/A",
            _interp(r_slope, r_r2),
            _fmt_float(t_slope, 2) if t_slope is not None else "N/A",
            _fmt_float(t_r2, 3) if t_r2 is not None else "N/A",
            _fmt_float(t_p, 4) if t_p is not None else "N/A",
            _interp(t_slope, t_r2),
            _r2_note(r_r2, t_r2, r_slope, t_slope),
        ]

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Highlight baris jika ada potensi leak (R²>=0.5)
        if (r_r2 is not None and r_r2 >= 0.5) or (t_r2 is not None and t_r2 >= 0.5):
            for ci in range(1, len(row_vals) + 1):
                ws.cell(row=row, column=ci).fill = PatternFill("solid", fgColor="FFFFF3E0")

        row += 1

    # Catatan metodologis
    row += 1
    ws.cell(row=row, column=1, value=(
        "Catatan: Soak test N=1, bersifat observasional — tidak ada uji inferensial. "
        "Fokus interpretasi pada R² (kekuatan trend) dan slope (MB/hour), bukan p-value. "
        "p selalu ≈0 untuk N ribuan sampel meskipun trend lemah."
    )).font = Font(italic=True, size=9, color="444444")
    ws.merge_cells(f"A{row}:J{row}")

    # Widths
    ws.column_dimensions["A"].width = 16
    for i in range(2, 11):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["I"].width = 28
    ws.column_dimensions["J"].width = 40
    ws.freeze_panes = "B4"


def write_order_effects_sheet(wb, all_results: dict):
    """
    OrderEffects sheet: verifikasi order effect dari counterbalancing 5:5.

    BUG-01 FIX: Fungsi lama (write_bonferroni_order_effects_sheet) menulis sheet
    dua kali — pass pertama (descriptive labels) ditimpa pass kedua (H-labels)
    karena reset row=3 di tengah fungsi yang sama. Fungsi baru ini hanya
    menulis order effects (Bonferroni dihapus dari pipeline).
    """
    ws = wb.create_sheet("OrderEffects")

    ws.merge_cells("A1:J1")
    ws["A1"] = "ORDER EFFECT VERIFICATION — Counterbalancing 5:5 REST-first/tRPC-first"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=COLORS["header_bg"])
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    ws.cell(row=row, column=1, value=(
        "Verifikasi bahwa Δ (REST−tRPC) konsisten antara rest-first dan trpc-first subgroup. "
        "Jika berlawanan arah atau magnitude sangat berbeda → order effect."
    )).font = Font(italic=True, size=9)
    ws.merge_cells(f"A{row}:J{row}")
    row += 2

    oe_headers = ["Scenario", "Metrik", "N REST-first", "N tRPC-first",
                  "Δ RF (REST−tRPC)", "Δ TF (REST−tRPC)", "Sama Arah?",
                  "Ratio Mag", "Flag", "Status"]
    for ci, h in enumerate(oe_headers, 1):
        _header_style(ws, row, ci, h, bg="FF1565C0", fg="FFFFFFFF")
    row += 1

    order_effects = all_results.get("order_effects", {})
    for gk in sorted(order_effects.keys()):
        gr     = order_effects[gk]
        sc     = gr.get("scenario", "?")
        n_rf   = gr.get("n_rest_first", 0)
        n_tf   = gr.get("n_trpc_first", 0)
        for mn, mdata in gr.get("metrics", {}).items():
            rf    = mdata.get("rest_first", {})
            tf    = mdata.get("trpc_first", {})
            flag  = mdata.get("order_effect_flag", False)
            same  = mdata.get("same_direction", True)
            ratio = mdata.get("magnitude_ratio")
            row_vals = [
                SCENARIO_LABELS.get(sc, sc),
                METRIC_LABELS.get(mn, mn),
                n_rf, n_tf,
                _fmt_float(rf.get("delta"), 3),
                _fmt_float(tf.get("delta"), 3),
                "Ya" if same else "⚠ TIDAK",
                _fmt_float(ratio, 2) if ratio is not None else "N/A",
                "⚠ FLAGGED" if flag else "OK",
                gr.get("note", "")[:40],
            ]
            for ci, val in enumerate(row_vals, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            if flag:
                for ci in range(1, len(row_vals) + 1):
                    ws.cell(row=row, column=ci).fill = PatternFill("solid", fgColor="FFFFF3E0")
            row += 1

    if not order_effects:
        ws.cell(row=row, column=1,
                value="(Data order effects tidak tersedia — run ulang dengan analyze.py yang sudah difix)")
        row += 1

    for ci in range(1, 11):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["J"].width = 40
    ws.freeze_panes = "A6"   # rows 1-5 frozen: title + description + blank + headers


# ---------------------------------------------------------------------------
# CHARTS
# ---------------------------------------------------------------------------

def plot_comparison_bar(scenario, test_type, metric_names, group, output_dir, condition="C2"):
    metrics_to_plot = [m for m in metric_names if m in group["metrics"]
                       and group["metrics"][m].get("status") not in ("no_data","insufficient_data")]
    if not metrics_to_plot:
        return None

    n_m   = len(metrics_to_plot)
    cols  = min(3, n_m)
    rows_ = math.ceil(n_m / cols)
    fig, axes = plt.subplots(rows_, cols, figsize=(5 * cols, 4 * rows_))
    if n_m == 1:
        axes = [[axes]]
    elif rows_ == 1:
        axes = [axes]
    axes = [ax for row in axes for ax in (row if hasattr(row, "__iter__") else [row])]

    cond_label = f" / {condition}" if condition != "C2" else ""
    fig.suptitle(f"{scenario.upper()} — {test_type.upper()}{cond_label} (N={group['n']})\nREST vs tRPC",
                 fontsize=13, fontweight="bold", y=1.01)

    for i, mn in enumerate(metrics_to_plot):
        ax   = axes[i]
        a    = group["metrics"][mn]
        desc = a.get("descriptive", {})
        rm, rs = desc.get("rest", {}).get("mean", 0), desc.get("rest", {}).get("std", 0)
        tm, ts = desc.get("trpc", {}).get("mean", 0), desc.get("trpc", {}).get("std", 0)
        rm, rs, tm, ts = rm or 0, rs or 0, tm or 0, ts or 0

        bars = ax.bar(["REST","tRPC"], [rm, tm], yerr=[rs, ts],
                      color=[COLORS["rest"], COLORS["trpc"]], alpha=0.85, capsize=5, width=0.5)
        ax.set_title(METRIC_LABELS.get(mn, mn), fontsize=10)
        ax.set_ylabel("Value", fontsize=8)
        for bar, val in zip(bars, [rm, tm]):
            ax.annotate(f"{val:.2f}", xy=(bar.get_x() + bar.get_width()/2, bar.get_height()),
                        xytext=(0, 3), textcoords="offset points", ha="center", va="bottom", fontsize=8)
        cd  = a.get("cohens_d", {})
        sig = a.get("inferential", {}).get("significant")
        if cd.get("d") is not None:
            cd_text = f"d={cd['d']:.2f} ({cd.get('magnitude','?')})" + (" *" if sig else "")
            ax.set_xlabel(cd_text, fontsize=8, color="darkgreen" if sig else "gray")
        ax.grid(axis="y", alpha=0.3)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    for j in range(n_m, len(axes)):
        axes[j].set_visible(False)
    plt.tight_layout()
    return _save(fig, output_dir, f"{scenario}_{test_type}_{condition}_comparison.png")


def plot_cohens_d_chart(scenario, test_type, group, output_dir, condition="C2"):
    metrics, d_vals, bar_colors = [], [], []
    for mn in INFERENTIAL_METRICS:
        a = group["metrics"].get(mn, {})
        if a.get("status") in ("no_data","insufficient_data"):
            continue
        cd = a.get("cohens_d", {})
        if cd.get("d") is None:
            continue
        metrics.append(METRIC_LABELS.get(mn, mn))
        d_vals.append(cd["d"])
        mag = cd.get("magnitude", "trivial")
        bar_colors.append(
            {"besar":"#2E7D32","sedang":"#F57F17","kecil":"#1565C0","trivial":"#757575"}.get(mag,"#9E9E9E")
        )
    if not metrics:
        return None

    fig, ax = plt.subplots(figsize=(max(6, len(metrics)*1.2), 5))
    bars = ax.barh(metrics, d_vals, color=bar_colors, alpha=0.8)
    for thresh, label, style in [(0.2,"kecil","--"),(-0.2,"","--"),(0.5,"sedang","-"),
                                   (-0.5,"","-"),(0.8,"besar",":"),(-0.8,"",":")]:
        ax.axvline(x=thresh, color=COLORS["threshold"], linestyle=style, alpha=0.5, linewidth=1)
        if label:
            ax.text(thresh, len(metrics)-0.5, f" {label}", fontsize=7, color=COLORS["threshold"], va="top")
    ax.axvline(x=0, color="black", linewidth=1)
    for bar, val in zip(bars, d_vals):
        ax.text(val+(0.02 if val>=0 else -0.02), bar.get_y()+bar.get_height()/2,
                f"{val:.3f}", va="center", ha="left" if val>=0 else "right", fontsize=9)
    cond_label = f" / {condition}" if condition != "C2" else ""
    ax.set_xlabel("Cohen's d\n(positif = REST lebih tinggi, negatif = tRPC lebih tinggi)", fontsize=9)
    ax.set_title(f"{scenario.upper()} — {test_type.upper()}{cond_label} — Effect Size (Cohen's d)",
                 fontsize=11, fontweight="bold")
    ax.grid(axis="x", alpha=0.3)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    plt.tight_layout()
    return _save(fig, output_dir, f"{scenario}_{test_type}_{condition}_cohens_d.png")


def plot_bootstrap_ci(scenario, test_type, group, output_dir, condition="C2"):
    metrics, ci_lowers, ci_uppers, means = [], [], [], []
    for mn in INFERENTIAL_METRICS:
        a = group["metrics"].get(mn, {})
        if a.get("status") in ("no_data","insufficient_data"):
            continue
        bci = a.get("bootstrap_ci", {})
        if bci.get("ci_lower") is None:
            continue
        metrics.append(METRIC_LABELS.get(mn, mn))
        ci_lowers.append(bci["ci_lower"])
        ci_uppers.append(bci["ci_upper"])
        means.append(bci.get("mean_d", 0))
    if not metrics:
        return None

    fig, ax = plt.subplots(figsize=(8, max(4, len(metrics)*0.8)))
    y_pos = range(len(metrics))
    xerr_low  = [m-lo for m,lo in zip(means,ci_lowers)]
    xerr_high = [hi-m  for m,hi in zip(means,ci_uppers)]
    ax.errorbar(means, list(y_pos), xerr=[xerr_low,xerr_high],
                fmt="o", color="#1565C0", capsize=5, markersize=7, linewidth=2)
    for i,(m,lo,hi) in enumerate(zip(means,ci_lowers,ci_uppers)):
        ax.plot(m, i, "o", color="#9E9E9E" if lo<=0<=hi else "#2E7D32", markersize=8, zorder=5)
    ax.axvline(x=0, color="red", linestyle="--", linewidth=1.5, alpha=0.7)
    ax.set_yticks(list(y_pos))
    ax.set_yticklabels(metrics, fontsize=9)
    ax.set_xlabel("Mean Difference REST − tRPC\n(dengan 95% Bootstrap CI)", fontsize=9)
    ax.set_title(f"{scenario.upper()} — {test_type.upper()} — Bootstrap 95% CI",
                 fontsize=11, fontweight="bold")
    ax.legend(handles=[
        mpatches.Patch(color="#2E7D32", label="CI tidak mencakup 0"),
        mpatches.Patch(color="#9E9E9E", label="CI mencakup 0"),
    ], fontsize=8, loc="lower right")
    ax.grid(axis="x", alpha=0.3)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    plt.tight_layout()
    return _save(fig, output_dir, f"{scenario}_{test_type}_bootstrap_ci.png")


def plot_boxplot(scenario, test_type, group, output_dir, condition="C2"):
    """BUG-02 FIX: Pakai actual vals, bukan approximation."""
    plot_data = []
    for mn in INFERENTIAL_METRICS:
        a = group["metrics"].get(mn, {})
        if a.get("status") in ("no_data","insufficient_data"):
            continue
        rv, tv = _get_chart_vals(a)
        if rv is None:
            continue
        plot_data.append((METRIC_LABELS.get(mn, mn), rv, tv))
    if not plot_data:
        return None

    n_m   = len(plot_data)
    cols  = min(3, n_m)
    rows_ = math.ceil(n_m/cols)
    fig, axes = plt.subplots(rows_, cols, figsize=(5*cols, 4*rows_))
    if n_m==1: axes=[[axes]]
    elif rows_==1: axes=[axes]
    axes=[ax for row in axes for ax in (row if hasattr(row,"__iter__") else [row])]

    cond_label = f" / {condition}" if condition != "C2" else ""
    fig.suptitle(f"{scenario.upper()} — {test_type.upper()}{cond_label} (N={group['n']})\nBoxplot REST vs tRPC",
                 fontsize=13, fontweight="bold", y=1.01)

    for i,(label,rv,tv) in enumerate(plot_data):
        ax = axes[i]
        bp = ax.boxplot([rv,tv], labels=["REST","tRPC"], patch_artist=True,
                        medianprops=dict(color="black", linewidth=2))
        bp["boxes"][0].set_facecolor(COLORS["rest"])
        bp["boxes"][1].set_facecolor(COLORS["trpc"])
        for box in bp["boxes"]: box.set_alpha(0.75)
        ax.set_title(label, fontsize=10)
        ax.set_ylabel("Value", fontsize=8)
        ax.grid(axis="y", alpha=0.3)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    for j in range(n_m, len(axes)): axes[j].set_visible(False)
    plt.tight_layout()
    return _save(fig, output_dir, f"{scenario}_{test_type}_{condition}_boxplot.png")


def plot_paired_scatter(scenario, test_type, group, output_dir, condition="C2"):
    """BUG-02 FIX: Pakai actual vals."""
    plot_data = []
    for mn in INFERENTIAL_METRICS:
        a = group["metrics"].get(mn, {})
        if a.get("status") in ("no_data","insufficient_data"):
            continue
        rv, tv = _get_chart_vals(a)
        if rv is None:
            continue
        plot_data.append((METRIC_LABELS.get(mn, mn), rv, tv))
    if not plot_data:
        return None

    n_m   = len(plot_data)
    cols  = min(3, n_m)
    rows_ = math.ceil(n_m/cols)
    fig, axes = plt.subplots(rows_, cols, figsize=(5*cols, 4.5*rows_))
    if n_m==1: axes=[[axes]]
    elif rows_==1: axes=[axes]
    axes=[ax for row in axes for ax in (row if hasattr(row,"__iter__") else [row])]

    cond_label = f" / {condition}" if condition != "C2" else ""
    fig.suptitle(f"{scenario.upper()} — {test_type.upper()}{cond_label}\nPaired Scatter (titik = 1 run)",
                 fontsize=13, fontweight="bold", y=1.01)

    for i,(label,rv,tv) in enumerate(plot_data):
        ax = axes[i]
        ax.scatter(rv, tv, color=COLORS["rest"], s=60, alpha=0.8, zorder=5)
        for idx,(x,y) in enumerate(zip(rv,tv)):
            ax.annotate(str(idx+1),(x,y),textcoords="offset points",xytext=(4,3),fontsize=7,color="gray")
        all_vals = rv+tv
        vmin, vmax = min(all_vals)*0.95, max(all_vals)*1.05
        ax.plot([vmin,vmax],[vmin,vmax],"k--",alpha=0.4,linewidth=1)
        ax.set_xlabel("REST",fontsize=8); ax.set_ylabel("tRPC",fontsize=8)
        ax.set_title(label,fontsize=10)
        ax.grid(alpha=0.3)
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
        above=sum(1 for x,y in zip(rv,tv) if y>x)
        below=sum(1 for x,y in zip(rv,tv) if y<x)
        ax.text(0.05,0.95,f"tRPC>REST: {above}\nREST>tRPC: {below}",
                transform=ax.transAxes,fontsize=7,va="top",
                bbox=dict(boxstyle="round,pad=0.3",facecolor="lightyellow",alpha=0.8))

    for j in range(n_m, len(axes)): axes[j].set_visible(False)
    plt.tight_layout()
    return _save(fig, output_dir, f"{scenario}_{test_type}_{condition}_paired_scatter.png")


def plot_dot_plot(scenario, test_type, group, output_dir, condition="C2"):
    """BUG-02 FIX: Pakai actual vals."""
    plot_data = []
    for mn in INFERENTIAL_METRICS:
        a = group["metrics"].get(mn, {})
        if a.get("status") in ("no_data","insufficient_data"):
            continue
        rv, tv = _get_chart_vals(a)
        if rv is None:
            continue
        plot_data.append((METRIC_LABELS.get(mn, mn), rv, tv, a.get("cohens_d", {})))
    if not plot_data:
        return None

    n_m   = len(plot_data)
    cols  = min(3, n_m)
    rows_ = math.ceil(n_m/cols)
    fig, axes = plt.subplots(rows_, cols, figsize=(4.5*cols, 4*rows_))
    if n_m==1: axes=[[axes]]
    elif rows_==1: axes=[axes]
    axes=[ax for row in axes for ax in (row if hasattr(row,"__iter__") else [row])]

    cond_label = f" / {condition}" if condition != "C2" else ""
    fig.suptitle(f"{scenario.upper()} — {test_type.upper()}{cond_label} (N={group['n']})\n"
                 f"Dot Plot — Titik Aktual per Run (Exploratory)",
                 fontsize=12, fontweight="bold", y=1.01)

    for i,(label,rv,tv,cd_info) in enumerate(plot_data):
        ax  = axes[i]
        n   = len(rv)
        jit = [-0.08+0.08*j for j in range(n)]
        ax.scatter([0+j for j in jit], rv, color=COLORS["rest"], s=80, alpha=0.9, zorder=5)
        ax.scatter([1+j for j in jit], tv, color=COLORS["trpc"], s=80, alpha=0.9, zorder=5)
        ax.hlines(sum(rv)/len(rv),-0.3,0.3,colors=COLORS["rest"],linewidths=2,linestyles="--",alpha=0.6)
        ax.hlines(sum(tv)/len(tv), 0.7,1.3,colors=COLORS["trpc"],linewidths=2,linestyles="--",alpha=0.6)
        ax.set_xticks([0,1]); ax.set_xticklabels(["REST","tRPC"],fontsize=9)
        ax.set_xlim(-0.5,1.5); ax.set_title(label,fontsize=10); ax.set_ylabel("Value",fontsize=8)
        ax.grid(axis="y",alpha=0.3); ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
        if cd_info.get("d") is not None:
            ax.set_xlabel(f"d={cd_info['d']:.2f} ({cd_info.get('magnitude','?')}) [exploratory]",
                          fontsize=7, color="gray")

    for j in range(n_m, len(axes)): axes[j].set_visible(False)
    plt.tight_layout()
    return _save(fig, output_dir, f"{scenario}_{test_type}_{condition}_dotplot.png")


def plot_soak_timeseries(scenario, group, output_dir, condition="C2"):
    """
    BUG-04-SOAK FIX: Setiap side pakai local variable x (tidak override).
    Tambah note bahwa REST dan tRPC berjalan independen (berbeda waktu).
    """
    ts      = group.get("timeseries") or {}
    rest_ts = ts.get("rest") or {}
    trpc_ts = ts.get("trpc") or {}
    if not rest_ts and not trpc_ts:
        return None

    def moving_avg(vals, window=10):
        result = []
        for i in range(len(vals)):
            start = max(0, i-window+1)
            result.append(sum(vals[start:i+1])/(i-start+1))
        return result

    panels = []
    for key, label, unit in [
        ("cpu_pct",      "CPU % (backend)",  "%"),
        ("mem_mb",       "RAM MB (backend)", "MB"),
        ("pg_active",    "DB Connections",   "conn"),
        ("network_kb_s", "Network I/O",      "KB/s"),
    ]:
        rv = rest_ts.get(key, [])
        tv = trpc_ts.get(key, [])
        if rv or tv:
            panels.append((key, label, unit, rv, tv))
    if not panels:
        return None

    n_panels = len(panels)
    fig, axes = plt.subplots(n_panels, 1, figsize=(14, 4*n_panels))
    if n_panels == 1:
        axes = [axes]

    fig.suptitle(
        f"{scenario.upper()} — SOAK TEST\n"
        "Time-Series (REST biru vs tRPC oranye, garis tebal = moving avg)\n"
        "Catatan: REST dan tRPC dijalankan independen (tidak bersamaan)",
        fontsize=12, fontweight="bold"
    )

    for ax, (key, label, unit, rv, tv) in zip(axes, panels):
        # BUG-04-SOAK FIX: Gunakan local variable per side (tidak override)
        if rv:
            x_rest = list(range(len(rv)))
            ax.plot(x_rest, rv, color=COLORS["rest"], alpha=0.25, linewidth=0.8)
            ax.plot(x_rest, moving_avg(rv), color=COLORS["ma_rest"], linewidth=2.0, label=f"REST (n={len(rv)})")

        if tv:
            x_trpc = list(range(len(tv)))
            ax.plot(x_trpc, tv, color=COLORS["trpc"], alpha=0.25, linewidth=0.8)
            ax.plot(x_trpc, moving_avg(tv), color=COLORS["ma_trpc"], linewidth=2.0, label=f"tRPC (n={len(tv)})")

        if key == "mem_mb":
            for side, ts_data, color in [("REST", rest_ts, COLORS["ma_rest"]),
                                          ("tRPC", trpc_ts, COLORS["ma_trpc"])]:
                slope_mh = ts_data.get("mem_slope_mb_per_hour")
                if slope_mh is not None:
                    ax.text(0.02, 0.95 if side=="REST" else 0.82,
                            f"{side}: {slope_mh:+.1f} MB/hr",
                            transform=ax.transAxes, fontsize=8, color=color, va="top",
                            bbox=dict(boxstyle="round,pad=0.2", facecolor="white", alpha=0.7))

        ax.set_ylabel(f"{label} ({unit})", fontsize=9)
        ax.legend(fontsize=8, loc="upper right")
        ax.grid(alpha=0.3)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    axes[-1].set_xlabel("Sample Index (lihat legend untuk jumlah sampel tiap sisi)", fontsize=9)
    plt.tight_layout()
    return _save(fig, output_dir, f"{scenario}_soak_timeseries.png")


def plot_percentile_profile(scenario, test_type, group, output_dir, condition="C2"):
    """ADD-02: Grouped bar P50/P90/P95/P99."""
    keys   = ["med_rt","p90","p95","p99"]
    labels = ["P50 (Median)","P90","P95","P99"]
    rest_v, trpc_v, valid_labels = [], [], []

    for key, label in zip(keys, labels):
        a    = group["metrics"].get(key, {})
        desc = a.get("descriptive", {})
        rm   = desc.get("rest", {}).get("mean")
        tm   = desc.get("trpc", {}).get("mean")
        if rm is not None and tm is not None:
            rest_v.append(rm); trpc_v.append(tm); valid_labels.append(label)

    if len(valid_labels) < 2:
        return None

    x = range(len(valid_labels))
    w = 0.35
    fig, ax = plt.subplots(figsize=(max(7, len(valid_labels)*2), 5))
    bars_r = ax.bar([xi-w/2 for xi in x], rest_v, w, label="REST", color=COLORS["rest"], alpha=0.85)
    bars_t = ax.bar([xi+w/2 for xi in x], trpc_v, w, label="tRPC", color=COLORS["trpc"], alpha=0.85)
    for bar, val in zip(bars_r, rest_v):
        ax.annotate(f"{val:.1f}", xy=(bar.get_x()+bar.get_width()/2, bar.get_height()),
                    xytext=(0,3), textcoords="offset points", ha="center", fontsize=8)
    for bar, val in zip(bars_t, trpc_v):
        ax.annotate(f"{val:.1f}", xy=(bar.get_x()+bar.get_width()/2, bar.get_height()),
                    xytext=(0,3), textcoords="offset points", ha="center", fontsize=8)
    for xi, rv, tv, lbl in zip(x, rest_v, trpc_v, valid_labels):
        gap     = tv - rv
        gap_pct = gap/rv*100 if rv != 0 else 0
        color   = "darkred" if gap > 0 else "darkgreen"
        ax.annotate(f"Δ{gap:+.1f}\n({gap_pct:+.0f}%)", xy=(xi, max(rv,tv)),
                    xytext=(0,12), textcoords="offset points", ha="center",
                    fontsize=7, color=color, fontweight="bold")
    cond_label = f" / {condition}" if condition != "C2" else ""
    ax.set_xticks(list(x)); ax.set_xticklabels(valid_labels, fontsize=10)
    ax.set_ylabel("Latency (ms)", fontsize=10)
    ax.set_title(f"{scenario.upper()} — {test_type.upper()}{cond_label}\n"
                 f"Percentile Profile REST vs tRPC (N={group['n']})",
                 fontsize=12, fontweight="bold")
    ax.legend(fontsize=10); ax.grid(axis="y", alpha=0.3)
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    plt.tight_layout()
    return _save(fig, output_dir, f"{scenario}_{test_type}_{condition}_percentile_profile.png")


def plot_sla_error_chart(all_groups: dict, output_dir: str):
    """ADD-04: SLA breach & functional error per group."""
    labels, rest_sla, trpc_sla, rest_err, trpc_err = [], [], [], [], []

    for group_key, group in sorted(all_groups.items()):
        r_sla = _get_mean(group,"sla_breach","rest")
        t_sla = _get_mean(group,"sla_breach","trpc")
        r_err = _get_mean(group,"functional_error","rest")
        t_err = _get_mean(group,"functional_error","trpc")
        if r_sla is None and t_sla is None and r_err is None and t_err is None:
            continue

        sc  = group["scenario"].replace("s0","S").replace("_browse","Bro").replace("_shopping","Shop").replace("_checkout","Chk").replace("_auth","Auth").replace("_admin","Adm")
        tt  = group["test_type"][:4].upper()
        labels.append(f"{sc}\n{tt}")
        rest_sla.append((r_sla or 0)*100); trpc_sla.append((t_sla or 0)*100)
        rest_err.append((r_err or 0)*100); trpc_err.append((t_err or 0)*100)

    if not labels:
        return None

    n, x, w = len(labels), range(len(labels)), 0.35
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(max(10, n*1.8), 9), sharex=True)
    fig.suptitle("SLA Breach Rate & Functional Error Rate — REST vs tRPC (semua group)",
                 fontsize=13, fontweight="bold")

    ax1.bar([xi-w/2 for xi in x], rest_sla, w, label="REST", color=COLORS["rest"], alpha=0.85)
    ax1.bar([xi+w/2 for xi in x], trpc_sla, w, label="tRPC", color=COLORS["trpc"], alpha=0.85)
    ax1.axhline(5, color="red", linestyle="--", linewidth=1, alpha=0.7, label="Threshold 5%")
    ax1.set_ylabel("SLA Breach Rate (%)", fontsize=10); ax1.set_title("SLA Breach Rate", fontsize=11)
    ax1.legend(fontsize=9); ax1.grid(axis="y", alpha=0.3)
    ax1.spines["top"].set_visible(False); ax1.spines["right"].set_visible(False)

    ax2.bar([xi-w/2 for xi in x], rest_err, w, label="REST", color=COLORS["rest"], alpha=0.85)
    ax2.bar([xi+w/2 for xi in x], trpc_err, w, label="tRPC", color=COLORS["trpc"], alpha=0.85)
    ax2.axhline(1, color="red", linestyle="--", linewidth=1, alpha=0.7, label="Threshold 1%")
    ax2.set_ylabel("Functional Error Rate (%)", fontsize=10); ax2.set_title("Functional Error Rate", fontsize=11)
    ax2.set_xticks(list(x)); ax2.set_xticklabels(labels, fontsize=8)
    ax2.legend(fontsize=9); ax2.grid(axis="y", alpha=0.3)
    ax2.spines["top"].set_visible(False); ax2.spines["right"].set_visible(False)

    plt.tight_layout()
    return _save(fig, output_dir, "ALL_sla_error_chart.png")


def plot_forest_plot(scenario, test_type, group, output_dir, condition="C2"):
    """
    ADD-05: Forest plot.
    ISS-04 FIX: Hapus CI fallback ±0.3 arbitrary. Jika CI tidak tersedia
    (SD=0 atau bootstrap gagal), tampilkan titik saja tanpa error bar.
    """
    all_metrics = INFERENTIAL_METRICS + ["db_query_avg_ms","network_total_kb_s","pg_active","sla_breach","functional_error"]
    rows = []
    for mn in all_metrics:
        a   = group["metrics"].get(mn, {})
        if a.get("status") in ("no_data","insufficient_data"):
            continue
        cd  = a.get("cohens_d", {})
        bci = a.get("bootstrap_ci", {})
        if cd.get("d") is None:
            continue

        d_val = cd["d"]
        desc  = a.get("descriptive", {})
        diff_sd = desc.get("diff", {}).get("std")

        # ISS-04 FIX: Hitung CI dalam Cohen's d space jika data cukup
        has_ci = False
        ci_lo_d, ci_hi_d = None, None
        if diff_sd and diff_sd > 0 and bci.get("ci_lower") is not None:
            ci_lo_d = bci["ci_lower"] / diff_sd
            ci_hi_d = bci["ci_upper"] / diff_sd
            has_ci  = True
        # Jika tidak ada CI yang valid, has_ci=False → plot titik saja

        rows.append({
            "label":     METRIC_LABELS.get(mn, mn),
            "d":         d_val,
            "ci_lo":     ci_lo_d,
            "ci_hi":     ci_hi_d,
            "has_ci":    has_ci,
            "magnitude": cd.get("magnitude","trivial"),
            "sig":       a.get("inferential", {}).get("significant"),
        })

    if not rows:
        return None

    fig, ax = plt.subplots(figsize=(10, max(5, len(rows)*0.6+2)))

    for i, row_ in enumerate(rows):
        color = {"besar":"#2E7D32","sedang":"#F57F17","kecil":"#1565C0","trivial":"#9E9E9E"}.get(row_["magnitude"],"#9E9E9E")

        if row_["has_ci"]:
            xerr_lo = row_["d"] - row_["ci_lo"]
            xerr_hi = row_["ci_hi"] - row_["d"]
            ax.errorbar(row_["d"], i, xerr=[[abs(xerr_lo)],[abs(xerr_hi)]],
                        fmt="o", color=color, capsize=4, markersize=8 if row_["sig"] else 6,
                        linewidth=1.5,
                        markeredgewidth=2 if row_["sig"] else 1,
                        markerfacecolor=color if row_["sig"] else "white",
                        markeredgecolor=color)
        else:
            # ISS-04 FIX: Titik saja tanpa error bar jika CI tidak tersedia
            ax.plot(row_["d"], i, "o",
                    color=color, markersize=8 if row_["sig"] else 6,
                    markerfacecolor=color if row_["sig"] else "white",
                    markeredgecolor=color, markeredgewidth=2)

    for thresh, label, style in [(0.2,"kecil","--"),(-0.2,"","--"),(0.5,"sedang","-"),(-0.5,"","-"),(0.8,"besar",":"),(-0.8,"",":")]:
        ax.axvline(x=thresh, color=COLORS["threshold"], linestyle=style, alpha=0.4, linewidth=1)
        if label:
            ax.text(thresh, len(rows)-0.3, f" {label}", fontsize=7, color=COLORS["threshold"], va="top")
    ax.axvline(x=0, color="black", linewidth=1.5)

    ax.set_yticks(range(len(rows)))
    ax.set_yticklabels([r["label"] for r in rows], fontsize=9)
    ax.set_xlabel("Cohen's d  (positif = REST lebih tinggi, negatif = tRPC lebih tinggi)\n"
                  "Solid = signifikan, kosong = tidak signifikan | error bar = Bootstrap CI (/SD)", fontsize=9)
    cond_label = f" / {condition}" if condition != "C2" else ""
    ax.set_title(f"{scenario.upper()} — {test_type.upper()}{cond_label}\n"
                 f"Forest Plot — Effect Size (N={group['n']})", fontsize=11, fontweight="bold")
    ax.legend(handles=[
        mpatches.Patch(color="#2E7D32", label="Besar (|d|≥0.8)"),
        mpatches.Patch(color="#F57F17", label="Sedang (0.5≤|d|<0.8)"),
        mpatches.Patch(color="#1565C0", label="Kecil (0.2≤|d|<0.5)"),
        mpatches.Patch(color="#9E9E9E", label="Trivial (|d|<0.2)"),
    ], fontsize=8, loc="lower right")
    ax.grid(axis="x", alpha=0.3)
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    plt.tight_layout()
    return _save(fig, output_dir, f"{scenario}_{test_type}_{condition}_forest_plot.png")


def write_chart_desc_sheet(wb, interpretations: dict):
    """
    Sheet 'ChartDesc' — deskripsi otomatis semua chart.
    Berguna sebagai basis caption gambar di skripsi.
    """
    ws = wb.create_sheet("ChartDesc")
    ws.merge_cells("A1:E1")
    ws["A1"] = "DESKRIPSI CHART — Basis Caption Gambar Skripsi"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=COLORS["header_bg"])
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    for ci, h in enumerate(["Filename","Chart Type","Caption Siap Pakai","Deskripsi Lengkap","Nomor Gambar (isi manual)"], 1):
        _header_style(ws, row, ci, h, bg=COLORS["header_bg"], fg=COLORS["header_fg"])
    row += 1

    chart_desc = interpretations.get("chart_descriptions", {})
    # Group by chart_type for organized display
    TYPE_ORDER = ["sla_error","bar_comparison","cohens_d","bootstrap_ci","boxplot",
                  "paired_scatter","percentile_profile","forest_plot","dotplot","soak_timeseries"]
    TYPE_COLORS = {
        "sla_error":        "FFE3F2FD",
        "bar_comparison":   "FFE8F5E9",
        "cohens_d":         "FFFFF9C4",
        "bootstrap_ci":     "FFE1F5FE",
        "boxplot":          "FFF3E5F5",
        "paired_scatter":   "FFF1F8E9",
        "percentile_profile":"FFFBE9E7",
        "forest_plot":      "FFE0F2F1",
        "dotplot":          "FFFFE8E1",
        "soak_timeseries":  "FFFFF3E0",
    }

    # Sort filenames by type then name
    sorted_items = sorted(
        chart_desc.items(),
        key=lambda x: (TYPE_ORDER.index(x[1].get("chart_type","")) if x[1].get("chart_type","") in TYPE_ORDER else 99, x[0])
    )

    for fname, info in sorted_items:
        ct  = info.get("chart_type","")
        bg  = TYPE_COLORS.get(ct, "FFFFFFFF")
        row_vals = [
            fname,
            ct,
            info.get("caption","")[:120],
            info.get("description","")[:300],
            "",  # nomor gambar diisi manual
        ]
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill      = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[row].height = 55
        row += 1

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# MAIN REPORT FUNCTION
# ---------------------------------------------------------------------------

def generate_report(analysis_results: dict, output_dir: str,
                    interpretations: dict = None):
    os.makedirs(output_dir, exist_ok=True)
    charts_dir = os.path.join(output_dir, "charts")
    os.makedirs(charts_dir, exist_ok=True)

    print(f"\nGenerating report ke: {output_dir}\n")

    # ── Raw JSON ──────────────────────────────────────────────────────────
    # ISS-06 FIX: strip difference_scores dari endpoint_metrics juga
    # order_effects masuk via slim (top-level non-groups keys)
    raw_json_path = os.path.join(output_dir, "ZENIT_Analysis_Raw.json")
    with open(raw_json_path, "w", encoding="utf-8") as f:
        # Semua top-level key selain "groups" masuk langsung — termasuk order_effects
        slim = {k: v for k, v in analysis_results.items() if k != "groups"}
        slim["groups"] = {}
        for gk, gv in analysis_results.get("groups", {}).items():
            slim_group = {k: v for k, v in gv.items() if k not in ("timeseries",)}
            # LIM-06 FIX: simpan scalar soak slope ke JSON
            if gv.get("test_type") == "soak" and gv.get("timeseries"):
                _SLOPE_SCALAR_KEYS = (
                    "mem_slope_mb_per_hour", "mem_slope_r2", "mem_slope_p",
                    "mem_slope_per_sample", "sampling_interval_sec", "mem_slope_note",
                )
                slim_group["soak_slope_summary"] = {
                    side: {k: ts[k] for k in _SLOPE_SCALAR_KEYS if k in ts}
                    for side, ts in gv["timeseries"].items()
                    if ts
                }
            slim_group["metrics"] = {
                mn: {k: v for k, v in ma.items() if k not in ("difference_scores","rest_vals","trpc_vals")}
                for mn, ma in gv.get("metrics", {}).items()
            }
            slim_group["endpoint_metrics"] = {
                ep: {k: v for k, v in ea.items() if k not in ("difference_scores","rest_vals","trpc_vals")}
                for ep, ea in gv.get("endpoint_metrics", {}).items()
            }
            slim["groups"][gk] = slim_group

        # Tambahkan summary_stats untuk akses cepat tanpa harus parse groups
        slim["summary_stats"] = _build_summary_stats(analysis_results)

        # Interpretation layer — hanya jika digenerate
        if interpretations:
            slim_interp = {
                "overall_summary": interpretations.get("overall_summary"),
                "research_questions": {
                    rq_id: {
                        "question": rq_data.get("question"),
                        "summary":  rq_data.get("summary") or rq_data.get("overall"),
                        "rest_conditions": rq_data.get("rest_conditions"),
                        "trpc_conditions": rq_data.get("trpc_conditions"),
                        "primary_recommendation": rq_data.get("primary_recommendation"),
                        "recommendations": rq_data.get("recommendations"),
                        "rest_summary": rq_data.get("rest_summary"),
                        "trpc_summary": rq_data.get("trpc_summary"),
                    }
                    for rq_id, rq_data in interpretations.get("research_questions", {}).items()
                },
                "patterns": interpretations.get("patterns", []),
                "soak":     interpretations.get("soak", {}),
                "group_summaries": {
                    gk: {
                        "overall":  gi.get("overall"),
                        "sig_wins": gi.get("sig_wins"),
                        "flags":    gi.get("flags"),
                        "metrics": {
                            mn: {
                                "winner":      mi.get("winner"),
                                "delta_abs":   mi.get("delta_abs"),
                                "delta_pct":   mi.get("delta_pct"),
                                "cohens_d":    mi.get("cohens_d"),
                                "magnitude":   mi.get("magnitude"),
                                "significant": mi.get("significant"),
                                "sentence":    mi.get("sentence"),
                                "practical":   mi.get("practical"),
                            }
                            for mn, mi in gi.get("metrics", {}).items() if mi
                        },
                    }
                    for gk, gi in interpretations.get("groups", {}).items()
                },
                "chart_descriptions": {
                    fname: {
                        "chart_type":  info.get("chart_type"),
                        "caption":     info.get("caption"),
                        "description": info.get("description"),
                    }
                    for fname, info in interpretations.get("chart_descriptions", {}).items()
                },
            }
            slim["interpretation"] = slim_interp

        json.dump(slim, f, indent=2, ensure_ascii=False, default=str)
    print(f"  ✓ Raw JSON: {raw_json_path}")

    # ── Excel ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, analysis_results)
    write_cross_scenario_sheet(wb, analysis_results)
    write_effect_size_matrix_sheet(wb, analysis_results)
    write_soak_comparison_sheet(wb, analysis_results)
    write_order_effects_sheet(wb, analysis_results)

    if interpretations:
        write_chart_desc_sheet(wb, interpretations)
        print(f"  ✓ New sheets: CrossScenario, EffectSizeMatrix, SoakComparison, OrderEffects, ChartDesc")
    else:
        print(f"  ✓ New sheets: CrossScenario, EffectSizeMatrix, SoakComparison, OrderEffects")

    soak_groups    = {}
    decomp_results = analysis_results.get("decompositions", {})

    for group_key, group in sorted(analysis_results.get("groups", {}).items()):
        write_detail_sheet(wb, group_key, group)
        if group["test_type"] == "soak":
            soak_groups[group_key] = group

    for group_key, group in soak_groups.items():
        write_soak_sheet(wb, group_key, group)

    if decomp_results:
        write_decomposition_sheet(wb, decomp_results)

    excel_path = os.path.join(output_dir, "ZENIT_Analysis_Report.xlsx")
    wb.save(excel_path)
    print(f"  ✓ Excel: {excel_path}")

    # ── Charts ────────────────────────────────────────────────────────────
    chart_files = []

    p = plot_sla_error_chart(analysis_results.get("groups", {}), charts_dir)
    if p:
        chart_files.append(p); print(f"  ✓ SLA/Error chart: {os.path.basename(p)}")

    bar_metrics = INFERENTIAL_METRICS + ["sla_breach","functional_error","db_query_avg_ms"]

    for group_key, group in sorted(analysis_results.get("groups", {}).items()):
        scenario  = group["scenario"]
        test_type = group["test_type"]
        condition = group.get("condition", "C2")

        if test_type == "soak":
            p = plot_soak_timeseries(scenario, group, charts_dir, condition)
            if p:
                chart_files.append(p); print(f"  ✓ Soak timeseries: {os.path.basename(p)}")
            continue

        p = plot_comparison_bar(scenario, test_type, bar_metrics, group, charts_dir, condition)
        if p:
            chart_files.append(p); print(f"  ✓ Bar chart: {os.path.basename(p)}")

        p = plot_cohens_d_chart(scenario, test_type, group, charts_dir, condition)
        if p:
            chart_files.append(p); print(f"  ✓ Cohen's d: {os.path.basename(p)}")

        if test_type == "load":
            p = plot_bootstrap_ci(scenario, test_type, group, charts_dir, condition)
            if p:
                chart_files.append(p); print(f"  ✓ Bootstrap CI: {os.path.basename(p)}")

            p = plot_boxplot(scenario, test_type, group, charts_dir, condition)
            if p:
                chart_files.append(p); print(f"  ✓ Boxplot: {os.path.basename(p)}")

            p = plot_paired_scatter(scenario, test_type, group, charts_dir, condition)
            if p:
                chart_files.append(p); print(f"  ✓ Paired scatter: {os.path.basename(p)}")

            p = plot_percentile_profile(scenario, test_type, group, charts_dir, condition)
            if p:
                chart_files.append(p); print(f"  ✓ Percentile profile: {os.path.basename(p)}")

            p = plot_forest_plot(scenario, test_type, group, charts_dir, condition)
            if p:
                chart_files.append(p); print(f"  ✓ Forest plot: {os.path.basename(p)}")

        elif test_type in ("stress","spike"):
            p = plot_dot_plot(scenario, test_type, group, charts_dir, condition)
            if p:
                chart_files.append(p); print(f"  ✓ Dot plot: {os.path.basename(p)}")

            p = plot_forest_plot(scenario, test_type, group, charts_dir, condition)
            if p:
                chart_files.append(p); print(f"  ✓ Forest plot: {os.path.basename(p)}")

    print(f"\nReport selesai. {len(chart_files)} chart dihasilkan.")
    print(f"   Excel: {excel_path}")
    return excel_path, chart_files


def _build_summary_stats(analysis_results: dict) -> dict:
    """
    Build ringkasan angka-angka kunci untuk akses cepat di JSON.
    Berguna untuk script post-processing atau copy-paste ke dokumen.

    Struktur: { "load_c2": { scenario: { metric: {rest, trpc, d, sig} } } }
    """
    summary = {"load_c2": {}, "soak": {}, "order_effects_clean": {}}
    groups  = analysis_results.get("groups", {})

    # Load C2 key numbers
    for sc_name in SCENARIO_ORDER:
        gk    = f"{sc_name}__load__C2"
        group = groups.get(gk)
        if not group:
            continue
        summary["load_c2"][sc_name] = {}
        for mn in CROSS_METRICS + ["p99", "payload_bytes", "sla_breach", "db_query_avg_ms"]:
            a  = group["metrics"].get(mn, {})
            if a.get("status") in ("no_data", "insufficient_data"):
                continue
            desc  = a.get("descriptive", {})
            cd    = a.get("cohens_d", {})
            inf   = a.get("inferential", {})
            rm    = desc.get("rest",  {}).get("mean")
            tm    = desc.get("trpc",  {}).get("mean")
            summary["load_c2"][sc_name][mn] = {
                "rest_mean": round(rm, 4) if rm is not None else None,
                "trpc_mean": round(tm, 4) if tm is not None else None,
                "delta":     round(rm - tm, 4) if rm is not None and tm is not None else None,
                "cohens_d":  round(cd.get("d"), 4) if cd.get("d") is not None else None,
                "magnitude": cd.get("magnitude"),
                "sig":       inf.get("significant"),
                "p":         round(inf.get("p"), 4) if inf.get("p") is not None else None,
            }

    # Soak slopes
    for sc_name in SCENARIO_ORDER:
        gk    = f"{sc_name}__soak__C2"
        group = groups.get(gk)
        if not group:
            continue
        ts      = group.get("timeseries") or {}
        rest_ts = ts.get("rest") or {}
        trpc_ts = ts.get("trpc") or {}
        summary["soak"][sc_name] = {
            "rest": {
                "slope_mb_per_hour": rest_ts.get("mem_slope_mb_per_hour"),
                "r2":                rest_ts.get("mem_slope_r2"),
                "p":                 rest_ts.get("mem_slope_p"),
            },
            "trpc": {
                "slope_mb_per_hour": trpc_ts.get("mem_slope_mb_per_hour"),
                "r2":                trpc_ts.get("mem_slope_r2"),
                "p":                 trpc_ts.get("mem_slope_p"),
            },
        }

    # Order effects — hanya flagged
    oe = analysis_results.get("order_effects", {})
    for gk, gr in oe.items():
        if gr.get("order_effect_detected"):
            summary["order_effects_clean"][gk] = {
                "scenario":    gr.get("scenario"),
                "note":        gr.get("note"),
                "flagged_metrics": [
                    mn for mn, md in gr.get("metrics", {}).items()
                    if md.get("order_effect_flag")
                ],
            }

    return summary